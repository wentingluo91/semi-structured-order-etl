"""
Semi-Structured Excel Order Form ETL
====================================

A sanitized portfolio version of a real-world automation workflow.

The script:
1. Scans a folder recursively for Excel order forms.
2. Detects a supported worksheet template.
3. Extracts customer, order, delivery, and cost information.
4. Standardizes dates and identifiers.
5. Creates business-friendly order numbers.
6. Exports an analytics-ready CSV file.

All company names, network paths, proprietary template names, and internal
identifiers have been removed. Update TEMPLATE_CONFIGS and field mappings to
match your own sample files.
"""

from __future__ import annotations

import argparse
import logging
import re
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


LOGGER = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Template configuration
# ---------------------------------------------------------------------------

TEMPLATE_CONFIGS = [
    {
        "sheet_name": "Standard Order Form",
        "max_row": 100,
        "max_col": 12,
        "summary_value_index": 11,
        "secondary_value_index": 10,
        "left_label_value_index": 2,
        "finish_value_column": 1,
    },
    {
        "sheet_name": "Alternate Order Form",
        "max_row": 100,
        "max_col": 15,
        "summary_value_index": 13,
        "secondary_value_index": 12,
        "left_label_value_index": 4,
        "finish_value_column": 2,
    },
]


SUMMARY_FIELDS = {
    "total cost": "total_cost",
    "delivery fee": "delivery_fee",
    "tax": "tax",
    "sub-total": "subtotal",
    "subtotal": "subtotal",
    "discount": "discount",
    "other": "other_charges",
    "customer cost": "customer_cost",
    "list price": "list_price",
    "total items": "total_items",
    "order date": "order_date",
    "stock confirmation": "stock_confirmation",
    "invoice": "invoice_number",
    "main item quantity": "main_item_quantity",
    "accessory quantity": "accessory_quantity",
}

SECONDARY_FIELDS = {
    "main item cost": "main_item_cost",
    "accessory cost": "accessory_cost",
    "customer multiplier": "customer_multiplier",
}

LEFT_LABEL_FIELDS = {
    "project name": "project_name",
    "job name": "project_name",
    "pickup date": "pickup_date",
    "pick-up date": "pickup_date",
    "delivery needed": "delivery_needed",
}

INLINE_FIELDS = {
    "packaging type": "packaging_type",
    "payment method": "payment_method",
}


EMPTY_RESULT = {
    "customer_id": None,
    "company_name": None,
    "address_line_1": None,
    "address_line_2": None,
    "phone": None,
    "email": None,
    "contact_name": None,
    "project_name": None,
    "pickup_date": None,
    "delivery_needed": None,
    "order_date": None,
    "stock_confirmation": None,
    "invoice_number": None,
    "finish": None,
    "packaging_type": None,
    "payment_method": None,
    "customer_multiplier": None,
    "total_cost": None,
    "delivery_fee": None,
    "tax": None,
    "subtotal": None,
    "discount": None,
    "other_charges": None,
    "customer_cost": None,
    "list_price": None,
    "total_items": None,
    "main_item_quantity": None,
    "accessory_quantity": None,
    "main_item_cost": None,
    "accessory_cost": None,
}


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def normalize_identifier(value: Any) -> str:
    """Convert a value into a safe identifier component."""
    if value is None:
        return "UNKNOWN"

    normalized = re.sub(r"[^A-Z0-9_-]+", "", str(value).strip().upper())
    return normalized or "UNKNOWN"


def normalize_date_key(value: Any) -> str:
    """Convert a date-like value into YYYYMMDD for order-number generation."""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "00000000"
    return parsed.strftime("%Y%m%d")


def format_cell_value(value: Any) -> Any:
    """Convert Excel values into CSV-friendly values."""
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, float):
        return round(value, 2)

    if isinstance(value, int):
        return value

    text = str(value).strip()
    return text or None


def value_after_colon(value: Any) -> str | None:
    """Return the text appearing after a normal or full-width colon."""
    if value is None:
        return None

    text = str(value).strip().replace("：", ":")
    return text.split(":", 1)[1].strip() if ":" in text else text


def find_customer_column(dataframe: pd.DataFrame) -> str:
    """Find the most likely customer identifier column."""
    candidates = (
        "customer_id",
        "customer",
        "client_id",
        "client",
        "account_id",
        "company_name",
    )

    lower_map = {str(column).lower(): column for column in dataframe.columns}

    for candidate in candidates:
        if candidate in lower_map:
            return lower_map[candidate]

    for column in dataframe.columns:
        lowered = str(column).lower()
        if "customer" in lowered or "client" in lowered:
            return column

    raise KeyError("No customer identifier column was found.")


def generate_order_numbers(
    dataframe: pd.DataFrame,
    date_column: str = "order_date",
    customer_column: str | None = None,
    prefix: str = "SO",
) -> pd.DataFrame:
    """Create sequential order numbers grouped by date and customer."""
    if date_column not in dataframe.columns:
        raise KeyError(f"Missing required date column: {date_column}")

    customer_column = customer_column or find_customer_column(dataframe)

    date_keys = dataframe[date_column].apply(normalize_date_key)
    customer_keys = dataframe[customer_column].apply(normalize_identifier)

    sequence = (
        dataframe.groupby([date_keys, customer_keys], dropna=False)
        .cumcount()
        .add(1)
    )

    dataframe = dataframe.copy()
    dataframe["order_number"] = [
        f"{prefix}-{date_key}-{customer_key}-{number:04d}"
        for date_key, customer_key, number
        in zip(date_keys, customer_keys, sequence)
    ]
    return dataframe


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def extract_customer_block(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    column: int,
    result: dict[str, Any],
) -> None:
    """Extract customer contact information below the customer identifier."""
    for offset in range(1, 11):
        cell = worksheet.cell(row=start_row + offset, column=column)
        if cell.value is None:
            continue

        text = str(cell.value).strip()
        label = text.split(":", 1)[0]
        normalized_label = "".join(
            character for character in label.lower() if character.isalpha()
        )

        if normalized_label in {"phone", "tel", "telephone", "mobile", "cell"}:
            result["phone"] = value_after_colon(text)
            continue

        if normalized_label in {"email", "mail"}:
            email = None
            hyperlink = getattr(cell, "hyperlink", None)

            if hyperlink and isinstance(hyperlink.target, str):
                if hyperlink.target.lower().startswith("mailto:"):
                    email = hyperlink.target.split(":", 1)[1].strip()

            result["email"] = email or value_after_colon(text)
            continue

        if normalized_label in {"contact", "attn", "attention"}:
            result["contact_name"] = value_after_colon(text)
            continue

        if result["company_name"] is None:
            result["company_name"] = text
        elif result["address_line_1"] is None:
            result["address_line_1"] = text
        elif result["address_line_2"] is None:
            result["address_line_2"] = text


def extract_with_template(
    workbook: openpyxl.Workbook,
    template: dict[str, Any],
) -> dict[str, Any]:
    """Extract one order using a configured worksheet layout."""
    worksheet = workbook[template["sheet_name"]]
    result = deepcopy(EMPTY_RESULT)

    rows = worksheet.iter_rows(
        min_row=1,
        max_row=template["max_row"],
        max_col=template["max_col"],
    )

    for row in rows:
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue

            text = cell.value.strip()
            lowered = text.lower().replace("：", ":")

            # Example supported labels: "Customer ID:" or "Client:"
            if result["customer_id"] is None and any(
                label in lowered
                for label in ("customer id", "customer:", "client id", "client:")
            ):
                identifier = value_after_colon(text)
                if identifier and identifier != text:
                    result["customer_id"] = identifier
                    extract_customer_block(
                        worksheet,
                        start_row=cell.row,
                        column=cell.column,
                        result=result,
                    )
                    continue

            if cell.column == 1:
                for keyword, field_name in LEFT_LABEL_FIELDS.items():
                    if keyword in lowered:
                        value = row[template["left_label_value_index"]].value
                        result[field_name] = format_cell_value(value)

            if "select finish" in lowered and result["finish"] is None:
                finish_value = worksheet.cell(
                    row=cell.row + 1,
                    column=template["finish_value_column"],
                ).value
                result["finish"] = format_cell_value(finish_value)

            for keyword, field_name in SUMMARY_FIELDS.items():
                if keyword in lowered:
                    value = row[template["summary_value_index"]].value
                    result[field_name] = format_cell_value(value)

            for keyword, field_name in SECONDARY_FIELDS.items():
                if keyword in lowered:
                    value = row[template["secondary_value_index"]].value
                    result[field_name] = format_cell_value(value)

            for keyword, field_name in INLINE_FIELDS.items():
                if keyword not in lowered or result[field_name] is not None:
                    continue

                inline_value = value_after_colon(text)
                if inline_value and inline_value != text:
                    result[field_name] = inline_value
                    continue

                start_index = cell.column
                for next_cell in row[start_index:]:
                    if next_cell.value is not None and str(next_cell.value).strip():
                        result[field_name] = str(next_cell.value).strip()
                        break

    result["template_name"] = template["sheet_name"]
    return result


def extract_order(file_path: Path) -> dict[str, Any]:
    """Load an Excel file and extract data using the first matching template."""
    workbook = openpyxl.load_workbook(
        filename=file_path,
        data_only=True,
        read_only=False,
    )

    available_sheets = set(workbook.sheetnames)

    for template in TEMPLATE_CONFIGS:
        if template["sheet_name"] not in available_sheets:
            continue
        return extract_with_template(workbook, template)

    supported = ", ".join(template["sheet_name"] for template in TEMPLATE_CONFIGS)
    raise ValueError(
        f"No supported worksheet was found in {file_path.name}. "
        f"Expected one of: {supported}"
    )


# ---------------------------------------------------------------------------
# Batch processing and export
# ---------------------------------------------------------------------------

def discover_excel_files(input_folder: Path) -> list[Path]:
    """Find non-temporary .xlsx files recursively."""
    return sorted(
        path
        for path in input_folder.rglob("*.xlsx")
        if not path.name.startswith("~$")
    )


def process_folder(input_folder: Path, output_csv: Path) -> pd.DataFrame:
    """Extract all supported order forms and write one consolidated CSV."""
    records: list[dict[str, Any]] = []

    for file_path in discover_excel_files(input_folder):
        try:
            LOGGER.info("Processing %s", file_path.name)
            record = extract_order(file_path)

            # Only relative locations are exported; no private drive paths.
            record["source_file"] = file_path.name
            record["source_folder"] = str(
                file_path.parent.relative_to(input_folder)
            )
            records.append(record)

        except Exception as error:
            LOGGER.warning("Skipped %s: %s", file_path.name, error)

    if not records:
        raise RuntimeError(
            "No supported Excel order forms were successfully processed."
        )

    dataframe = pd.DataFrame(records)

    if "order_date" in dataframe.columns:
        try:
            dataframe = generate_order_numbers(dataframe)
        except (KeyError, ValueError) as error:
            LOGGER.warning("Order-number generation was skipped: %s", error)

        sort_key = pd.to_datetime(dataframe["order_date"], errors="coerce")
        dataframe = (
            dataframe.assign(_sort_key=sort_key)
            .sort_values("_sort_key", ascending=False, na_position="last")
            .drop(columns="_sort_key")
        )

    preferred_columns = [
        "source_file",
        "source_folder",
        "template_name",
        "order_number",
        "customer_id",
        "company_name",
        "project_name",
        "order_date",
        "pickup_date",
        "delivery_needed",
        "invoice_number",
        "stock_confirmation",
    ]

    ordered_columns = [
        column for column in preferred_columns if column in dataframe.columns
    ]
    ordered_columns.extend(
        column for column in dataframe.columns if column not in ordered_columns
    )
    dataframe = dataframe[ordered_columns]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_csv, index=False, encoding="utf-8-sig")
    LOGGER.info("Created %s", output_csv)

    return dataframe


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract structured records from semi-structured Excel order forms."
        )
    )
    parser.add_argument(
        "input_folder",
        type=Path,
        help="Folder containing sample .xlsx order forms.",
    )
    parser.add_argument(
        "output_csv",
        type=Path,
        help="Destination CSV file.",
    )
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    args = parse_arguments()

    if not args.input_folder.exists():
        raise FileNotFoundError(
            f"Input folder does not exist: {args.input_folder}"
        )

    process_folder(args.input_folder, args.output_csv)


if __name__ == "__main__":
    input_folder = Path(r"C:\Users\25214\Desktop\Profolio\sample_data")
    output_csv = Path(r"C:\Users\25214\Desktop\Profolio\orders.csv")

    process_folder(input_folder, output_csv)